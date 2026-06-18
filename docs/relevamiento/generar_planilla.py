#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Genera la planilla de relevamiento + rentabilidad + métricas de éxito para Lince.
Salida: Relevamiento-Lince.xlsx (mismo directorio).

Convención de colores:
  - Amarillo  = celda a COMPLETAR en el relevamiento.
  - Gris      = celda CALCULADA (fórmula); no tocar.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---- Paleta (marca Lince) ----
INK = "1B2B23"
MOSS = "3D5A45"
SAGE = "6E8579"
INPUT = "FFF7E6"   # amarillo suave: completar
CALC = "EFEAE0"    # gris: calculado

f_title = Font(bold=True, size=15, color=INK)
f_section = Font(bold=True, color="FFFFFF")
f_header = Font(bold=True, color="FFFFFF")
f_calc = Font(bold=True, color=INK)
f_note = Font(italic=True, color=SAGE, size=10)

fill_section = PatternFill("solid", fgColor=MOSS)
fill_header = PatternFill("solid", fgColor=SAGE)
fill_input = PatternFill("solid", fgColor=INPUT)
fill_calc = PatternFill("solid", fgColor=CALC)

thin = Side(style="thin", color="C9C4B8")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap = Alignment(wrap_text=True, vertical="top")
top = Alignment(vertical="top")

wb = Workbook()


def title_row(ws, text, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=text)
    c.font = f_title
    ws.row_dimensions[1].height = 24


def section(ws, r, text, ncols):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    c = ws.cell(row=r, column=1, value=text)
    c.font = f_section
    c.fill = fill_section
    c.alignment = top
    return r + 1


def headers(ws, r, cols):
    for i, h in enumerate(cols, start=1):
        c = ws.cell(row=r, column=i, value=h)
        c.font = f_header
        c.fill = fill_header
        c.alignment = wrap
        c.border = border
    return r + 1


def form_sheet(title, rows, widths=(36, 32, 52)):
    """rows: ('SEC', texto) | (campo, nota). Columna B = respuesta (amarillo)."""
    ws = wb.create_sheet(title)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    title_row(ws, title, len(widths))
    r = headers(ws, 2, ["Campo", "Respuesta", "Notas / aclaraciones"])
    for row in rows:
        if row[0] == "SEC":
            r = section(ws, r, row[1], len(widths))
            continue
        campo, nota = row
        ws.cell(row=r, column=1, value=campo).alignment = wrap
        ws.cell(row=r, column=1).border = border
        b = ws.cell(row=r, column=2)
        b.fill = fill_input
        b.border = border
        b.alignment = wrap
        n = ws.cell(row=r, column=3, value=nota)
        n.font = f_note
        n.alignment = wrap
        n.border = border
        ws.row_dimensions[r].height = 20
        r += 1
    ws.freeze_panes = "A3"
    return ws


# ============================ INSTRUCCIONES ============================
ws = wb.active
ws.title = "Instrucciones"
ws.column_dimensions["A"].width = 100
title_row(ws, "Relevamiento + Rentabilidad — Lince", 1)
intro = [
    "",
    "Cómo usar esta planilla:",
    "•  Completá SOLO las celdas amarillas (datos del relevamiento).",
    "•  Las celdas grises se calculan solas (fórmulas) — no las toques.",
    "•  Una pestaña por tema. Empezá por 'Negocio' y seguí el orden.",
    "•  Si la subís a Google Sheets, las fórmulas se mantienen.",
    "",
    "Pestañas:",
    "   1. Negocio                — datos generales",
    "   2. Operación y ventas     — canales, caja, delivery, WhatsApp",
    "   3. Productos y precios     — SKUs, listas de proveedores, actualización de precios",
    "   4. Stock y proveedores",
    "   5. Clientes               — tráfico, ticket, fiado, recurrencia",
    "   6. Finanzas y rentabilidad — ventas, margen, gastos, ganancia (con fórmulas)",
    "   7. Diagnóstico             — dolores detectados y servicio propuesto",
    "   8. Métricas de éxito       — KPIs antes/objetivo/actual + ROI del proyecto",
    "",
    "Objetivo: entender el negocio, estimar su rentabilidad y elegir UN caso piloto",
    "de alto impacto para el primer cliente (gratis), con métricas para medir el éxito.",
]
r = 3
for line in intro:
    c = ws.cell(row=r, column=1, value=line)
    c.alignment = wrap
    if line.endswith(":"):
        c.font = Font(bold=True, color=INK)
    r += 1
# meta
r += 1
for campo, nota in [("Fecha del relevamiento", ""), ("Relevado por", ""), ("Negocio", "")]:
    ws.cell(row=r, column=1, value=campo).font = Font(bold=True, color=INK)
    r += 1

# ============================ 1. NEGOCIO ============================
form_sheet("Negocio", [
    ("SEC", "Datos generales"),
    ("Nombre del negocio", ""),
    ("Rubro", "Bazar / Supermercado / Ambos"),
    ("¿Una sucursal o varias? ¿cuántas?", ""),
    ("Dirección(es)", ""),
    ("Dueño / contacto", "Nombre, teléfono"),
    ("Antigüedad del negocio", "Años"),
    ("Superficie aprox (m²)", ""),
    ("Cantidad de empleados", ""),
    ("Horario de atención", ""),
    ("Estacionalidad / picos", "Ej: fiestas, inicio de clases, fines de semana"),
    ("Nivel de digitalización", "Nada / planilla / software de gestión"),
    ("SEC", "Objetivo del dueño"),
    ("Objetivo principal", "Vender más / ahorrar tiempo / ordenar precios / reducir quiebres"),
    ("¿Qué tarea odia o repite más?", "LA pregunta de oro"),
    ("Expectativa con la automatización", ""),
])

# ====================== 2. OPERACIÓN Y VENTAS ======================
form_sheet("Operación y ventas", [
    ("SEC", "Cómo vende"),
    ("Modalidad", "Mostrador / autoservicio / ambos"),
    ("¿Hace delivery?", "Sí/No — zona de cobertura"),
    ("¿Cómo toma los pedidos hoy?", "Teléfono / WhatsApp / presencial"),
    ("Tiempo diario tomando/coordinando pedidos", "Horas aprox"),
    ("¿Tiene catálogo o tienda online?", "Sí/No — cuál"),
    ("SEC", "Caja y pagos"),
    ("Sistema de caja / POS", "¿Cuál? (software, balanza, planilla, nada)"),
    ("¿El POS exporta datos?", "Sí/No — formato (Excel/CSV)"),
    ("Medios de pago aceptados", "Efectivo, débito, crédito, MercadoPago, transferencia, QR"),
    ("SEC", "Canales digitales"),
    ("¿Usa WhatsApp para vender?", "Sí/No — ¿WhatsApp Business?"),
    ("Mensajes de clientes por día (aprox)", ""),
    ("¿Tiene catálogo de WhatsApp cargado?", "Sí/No"),
    ("Tiempo diario respondiendo consultas repetidas", "Horas aprox"),
    ("Redes (Instagram/Facebook)", "¿Activas? ¿quién las maneja?"),
    ("Ficha de Google Business", "Sí/No — ¿actualizada?"),
])

# ====================== 3. PRODUCTOS Y PRECIOS ======================
form_sheet("Productos y precios", [
    ("SEC", "Catálogo"),
    ("Cantidad aproximada de SKUs", "Productos distintos"),
    ("Categorías principales", ""),
    ("¿Tiene lista de precios maestra digital?", "Sí/No — dónde (Excel, sistema)"),
    ("SEC", "Precios e inflación"),
    ("¿Cómo llegan las listas de proveedores?", "PDF / Excel / foto / papel / web"),
    ("Frecuencia de cambios de precio", "Diaria / semanal / quincenal / mensual"),
    ("¿Cómo actualiza precios hoy?", "A mano / con sistema"),
    ("Tiempo por actualizar cada lista", "Horas aprox"),
    ("Regla de margen (markup)", "¿% uniforme o por categoría?"),
    ("Errores de precio detectados", "Frecuencia / impacto"),
    ("SEC", "Cartelería"),
    ("¿Usa etiquetas / carteles de góndola?", "Sí/No"),
    ("¿Cómo los hace hoy?", "A mano / Word / sistema"),
    ("Frecuencia de reimpresión de carteles", ""),
])

# ====================== 4. STOCK Y PROVEEDORES ======================
form_sheet("Stock y proveedores", [
    ("SEC", "Stock"),
    ("¿Controla stock?", "No / a ojo / planilla / sistema"),
    ("¿Sufre quiebres (faltantes)?", "Frecuencia"),
    ("¿Sobrestock / capital inmovilizado?", ""),
    ("Perecederos (super): ¿problema de vencimientos?", "Sí/No"),
    ("Desperdicio por vencimientos (aprox $/mes)", ""),
    ("SEC", "Proveedores"),
    ("Cantidad de proveedores principales", ""),
    ("Proveedores top", "Lista"),
    ("Condiciones de pago a proveedores", "Contado / cuenta corriente / plazos"),
    ("¿Compara precios entre proveedores?", "Sí/No — cómo"),
])

# ============================ 5. CLIENTES ============================
form_sheet("Clientes", [
    ("SEC", "Tráfico y consumo"),
    ("Clientes por día (aprox)", ""),
    ("Ticket promedio ($)", ""),
    ("% de clientes recurrentes (aprox)", ""),
    ("SEC", "Fiado / cuenta corriente"),
    ("¿Da fiado / cuenta corriente?", "Sí/No"),
    ("¿Cómo registra el fiado?", "Cuaderno / planilla / app"),
    ("Monto promedio de cuenta corriente", ""),
    ("Morosidad / demora de cobro", ""),
    ("SEC", "Base de clientes y fidelización"),
    ("¿Tiene base de clientes (teléfonos)?", "Sí/No — cantidad"),
    ("¿Hace promociones / fidelización?", "Sí/No — cómo"),
])

# ====================== 6. FINANZAS Y RENTABILIDAD ======================
ws = wb.create_sheet("Finanzas y rentabilidad")
for i, w in enumerate((44, 20, 50), start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
title_row(ws, "Finanzas y rentabilidad (mensual)", 3)
r = headers(ws, 2, ["Concepto", "Valor", "Notas"])


def fin_row(r, concepto, nota="", formula=None, money=True, pct=False, calc=False):
    ws.cell(row=r, column=1, value=concepto).alignment = wrap
    ws.cell(row=r, column=1).border = border
    b = ws.cell(row=r, column=2)
    b.border = border
    if formula is not None:
        b.value = formula
        b.fill = fill_calc
        b.font = f_calc
    else:
        b.fill = fill_input
    if pct:
        b.number_format = "0.0%"
    elif money:
        b.number_format = '#,##0'
    n = ws.cell(row=r, column=3, value=nota)
    n.font = f_note
    n.alignment = wrap
    n.border = border
    return r + 1


r = section(ws, r, "Ingresos y costo de mercadería", 3)
R_VENTAS = r;        r = fin_row(r, "Ventas mensuales ($)", "Facturación total del mes")
R_CMVPCT = r;        r = fin_row(r, "Costo de mercadería (% de ventas)", "Ej: 70% (lo que cuesta lo que vendés)", money=False, pct=True)
R_CMV = r;           r = fin_row(r, "Costo de mercadería ($)", "= Ventas × CMV%", formula=f"=B{R_VENTAS}*B{R_CMVPCT}", calc=True)
R_MB = r;            r = fin_row(r, "Margen bruto ($)", "= Ventas − CMV", formula=f"=B{R_VENTAS}-B{R_CMV}", calc=True)
R_MBPCT = r;         r = fin_row(r, "Margen bruto (%)", "= Margen bruto / Ventas", formula=f"=IFERROR(B{R_MB}/B{R_VENTAS},0)", money=False, pct=True, calc=True)

r = section(ws, r, "Gastos fijos mensuales", 3)
R_ALQ = r;           r = fin_row(r, "Alquiler ($)", "")
R_SUELDOS = r;       r = fin_row(r, "Sueldos ($)", "")
R_SERV = r;          r = fin_row(r, "Servicios (luz/agua/gas/internet) ($)", "")
R_IMP = r;           r = fin_row(r, "Impuestos y tasas ($)", "")
R_OTROS = r;         r = fin_row(r, "Otros gastos fijos ($)", "")
R_GF = r;            r = fin_row(r, "Total gastos fijos ($)", "= suma de arriba", formula=f"=SUM(B{R_ALQ}:B{R_OTROS})", calc=True)

r = section(ws, r, "Resultado", 3)
R_GO = r;            r = fin_row(r, "Ganancia operativa ($)", "= Margen bruto − Gastos fijos", formula=f"=B{R_MB}-B{R_GF}", calc=True)
R_MN = r;            r = fin_row(r, "Margen neto (%)", "= Ganancia / Ventas", formula=f"=IFERROR(B{R_GO}/B{R_VENTAS},0)", money=False, pct=True, calc=True)

r = section(ws, r, "Control cruzado (opcional)", 3)
R_TICKET = r;        r = fin_row(r, "Ticket promedio ($)", "")
R_CLI = r;           r = fin_row(r, "Clientes por día", "", money=False)
r = fin_row(r, "Ventas estimadas/mes (control)", "= ticket × clientes × 30", formula=f"=B{R_TICKET}*B{R_CLI}*30", calc=True)
ws.freeze_panes = "A3"

# ====================== 7. DIAGNÓSTICO ======================
ws = wb.create_sheet("Diagnóstico")
cols = ["Dolor / problema detectado", "Impacto (Alto/Medio/Bajo)", "Servicio Lince propuesto",
        "Esfuerzo (A/M/B)", "Prioridad (1-5)", "Métrica que mejoraría"]
for i, w in enumerate((40, 16, 34, 14, 12, 34), start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
title_row(ws, "Diagnóstico y oportunidades", len(cols))
r = headers(ws, 2, cols)
candidatos = [
    ("Actualizar precios desde listas de proveedores lleva mucho tiempo", "", "Automatización de precios + carteles (IA)", "", "", "Horas/lista; errores de precio"),
    ("Se pierden pedidos/consultas por WhatsApp", "", "Catálogo + pedidos por WhatsApp", "", "", "Pedidos/mes; tiempo de atención"),
    ("Faltantes / quiebres de stock", "", "Alertas de reposición", "", "", "Quiebres/mes"),
    ("Fiado desordenado / cobro lento", "", "Recordatorios de cuenta corriente", "", "", "% al día; días de cobro"),
    ("Clientes no recompran", "", "Pedido habitual / fidelización", "", "", "Tasa de recompra"),
    ("Desperdicio por vencimientos (super)", "", "Alertas de vencimiento + promo", "", "", "Desperdicio $/mes"),
    ("Poca visibilidad en Google", "", "Ficha de Google (baja prioridad)", "", "", "Llamados/visitas"),
]
for fila in candidatos:
    for i, val in enumerate(fila, start=1):
        c = ws.cell(row=r, column=i, value=val)
        c.alignment = wrap
        c.border = border
        if i != 1 and i != 3 and i != 6:
            c.fill = fill_input
        ws.row_dimensions[r].height = 30
    r += 1
for _ in range(5):  # filas en blanco para sumar más
    for i in range(1, len(cols) + 1):
        c = ws.cell(row=r, column=i)
        c.border = border
        c.fill = fill_input
    r += 1
ws.freeze_panes = "A3"

# ====================== 8. MÉTRICAS DE ÉXITO ======================
ws = wb.create_sheet("Métricas de éxito")
cols = ["Métrica (KPI)", "Cómo se mide / fuente", "Sentido", "Línea base (antes)",
        "Objetivo", "Actual", "Variación %", "Frecuencia"]
for i, w in enumerate((34, 30, 10, 16, 14, 14, 12, 16), start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
title_row(ws, "Métricas de éxito de la implementación", len(cols))
r = headers(ws, 2, cols)


def kpi(r, nombre, fuente, sentido, freq):
    vals = [nombre, fuente, sentido, None, None, None, None, freq]
    for i, val in enumerate(vals, start=1):
        c = ws.cell(row=r, column=i, value=val)
        c.alignment = wrap
        c.border = border
        if i in (4, 5, 6):           # base / objetivo / actual = completar
            c.fill = fill_input
        if i == 7:                    # variación = fórmula
            c.value = f'=IFERROR((F{r}-D{r})/D{r},"")'
            c.fill = fill_calc
            c.number_format = "0.0%"
    ws.row_dimensions[r].height = 28
    return r + 1


def kpi_section(r, txt):
    return section(ws, r, txt, len(cols))


r = kpi_section(r, "Eficiencia operativa")
r = kpi(r, "Horas/semana en tareas repetitivas", "Atención + precios + pedidos", "↓ mejor", "Semanal")
r = kpi(r, "Tiempo de actualizar precios por lista (min)", "Cronómetro", "↓ mejor", "Por lista")
r = kpi(r, "Errores de precio por mes", "Conteo", "↓ mejor", "Mensual")
r = kpi_section(r, "Comercial / ventas")
r = kpi(r, "Pedidos por WhatsApp / mes", "Bot / registro", "↑ mejor", "Mensual")
r = kpi(r, "Ventas mensuales ($)", "Caja / POS", "↑ mejor", "Mensual")
r = kpi(r, "Ticket promedio ($)", "Caja / POS", "↑ mejor", "Mensual")
r = kpi(r, "Clientes nuevos / mes", "Registro", "↑ mejor", "Mensual")
r = kpi(r, "Tasa de recompra (%)", "Registro de clientes", "↑ mejor", "Mensual")
r = kpi_section(r, "Atención")
r = kpi(r, "Consultas respondidas < 5 min (%)", "Bot", "↑ mejor", "Semanal")
r = kpi(r, "Consultas fuera de horario capturadas / mes", "Bot", "↑ mejor", "Mensual")
r = kpi_section(r, "Stock")
r = kpi(r, "Quiebres (faltantes) / mes", "Conteo", "↓ mejor", "Mensual")
r = kpi(r, "Desperdicio por vencimientos ($/mes)", "Estimación", "↓ mejor", "Mensual")
r = kpi_section(r, "Cobranzas (fiado)")
r = kpi(r, "Cuenta corriente al día (%)", "Planilla de fiado", "↑ mejor", "Mensual")
r = kpi(r, "Días promedio de cobro", "Planilla de fiado", "↓ mejor", "Mensual")
r = kpi_section(r, "Adopción / satisfacción")
r = kpi(r, "% de pedidos por el canal nuevo", "Bot / caja", "↑ mejor", "Mensual")
r = kpi(r, "Satisfacción del dueño (1-10)", "Encuesta", "↑ mejor", "Mensual")

# ---- ROI del proyecto (el éxito, en plata) ----
r += 1
r = section(ws, r, "ROI del proyecto (impacto económico)", len(cols))


def roi_row(r, concepto, nota="", formula=None, money=True, pct=False):
    ws.cell(row=r, column=1, value=concepto).alignment = wrap
    ws.cell(row=r, column=1).border = border
    ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)
    b = ws.cell(row=r, column=4)
    b.border = border
    if formula is not None:
        b.value = formula
        b.fill = fill_calc
        b.font = f_calc
    else:
        b.fill = fill_input
    b.number_format = "0.0%" if pct else ('#,##0' if money else "0.0")
    n = ws.cell(row=r, column=2, value=nota)
    n.font = f_note
    n.alignment = wrap
    return r + 1


R_HS = r;     r = roi_row(r, "Horas/semana ahorradas", "Tiempo que libera la automatización")
R_VH = r;     r = roi_row(r, "Valor de la hora ($)", "Costo de esa hora para el negocio")
R_AHORRO = r; r = roi_row(r, "Ahorro mensual ($)", "= horas × valor × 4,33", formula=f"=D{R_HS}*D{R_VH}*4.33")
R_PED = r;    r = roi_row(r, "Pedidos incrementales / mes", "Pedidos nuevos que captás", money=False)
R_TK = r;     r = roi_row(r, "Ticket promedio ($)", "")
R_VINC = r;   r = roi_row(r, "Ventas incrementales ($)", "= pedidos × ticket", formula=f"=D{R_PED}*D{R_TK}")
R_MGP = r;    r = roi_row(r, "Margen bruto (%)", "Traer de la pestaña Finanzas", pct=True)
R_GINC = r;   r = roi_row(r, "Ganancia incremental ($)", "= ventas incr. × margen", formula=f"=D{R_VINC}*D{R_MGP}")
R_OTRO = r;   r = roi_row(r, "Otros ahorros ($/mes)", "Menos quiebres / desperdicio")
R_COSTO = r;  r = roi_row(r, "Costo del servicio Lince ($/mes)", "Tu abono mensual")
R_BEN = r;    r = roi_row(r, "Beneficio neto mensual ($)", "= ahorro + ganancia + otros − costo", formula=f"=D{R_AHORRO}+D{R_GINC}+D{R_OTRO}-D{R_COSTO}")
R_ROI = r;    r = roi_row(r, "ROI mensual (%)", "= beneficio / costo", formula=f"=IFERROR(D{R_BEN}/D{R_COSTO},0)", pct=True)
R_SETUP = r;  r = roi_row(r, "Inversión inicial / setup ($)", "Cobro único de armado")
R_PAY = r;    r = roi_row(r, "Repago (meses)", "= setup / beneficio neto", formula=f"=IFERROR(D{R_SETUP}/D{R_BEN},0)", money=False)

ws.freeze_panes = "A3"

import os
out = os.path.join(os.path.dirname(__file__), "Relevamiento-Lince.xlsx")
wb.save(out)
print("Generado:", out)
print("Pestañas:", wb.sheetnames)
